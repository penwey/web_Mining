# -*- coding:utf-8 -*-
from openpyxl import Workbook



def file_create(file_name, file_list):
    wb = Workbook()
    ws = wb.active
    nrows = 1
    ws.cell(nrows, 1, 'job_title')
    ws.cell(nrows, 2, 'salary')
    ws.cell(nrows, 3, 'area')
    ws.cell(nrows, 4, 'education')
    ws.cell(nrows, 5, 'working_seniority')
    ws.cell(nrows, 6, 'company')
    ws.cell(nrows, 7, 'offer_time')
    ws.cell(nrows, 8, 'info_url')
    ws.cell(nrows, 9, 'create_time')
    for item in file_list:
        nrows += 1
        ws.cell(nrows, 1, item['job_title'])
        ws.cell(nrows, 2, item['salary'])
        ws.cell(nrows, 3, item['area'])
        ws.cell(nrows, 4, item['education'])
        ws.cell(nrows, 5, item['working_seniority'])
        ws.cell(nrows, 6, item['company'])
        ws.cell(nrows, 7, item['offer_time'])
        ws.cell(nrows, 8, item['info_url'])
        ws.cell(nrows, 9, item['create_time'])
        
    wb.save(file_name)
    print('数据写入成功!')

